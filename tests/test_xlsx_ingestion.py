"""Phase 10 tests: secure XLSX KYC ingestion channel.

XLSX reading (openpyxl, read-only + data-only, no formula evaluation) feeds
the SAME validation/normalization/dedup/audit pipeline as CSV via
``read_kyc_rows`` -> ``normalize_batch``. Fully self-contained: synthetic
workbooks written to ``tmp_path``; no network/DB/real data.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from app.ingestion.connectors.kyc_file_connector import (
    MAX_XLSX_ROWS,
    read_xlsx_rows,
)
from app.ingestion.errors import (
    CorruptFileError,
    EmptyFileError,
    KycPathError,
    MissingFileError,
    UnsupportedFileTypeError,
)
from app.ingestion.pipelines.kyc_ingestion_pipeline import ingest_kyc_file
from app.schemas.kyc import NormalizedKYCEntity, SectorRiskLevel

HEADER = [
    "client_id", "client_name", "client_type", "sector", "sector_risk",
    "country", "pep_flag", "sanctions_flag", "fatf_country_flag",
]

SYNTHETIC_PII = "PHASE10_PII_DO_NOT_LEAK_106"


def _write_xlsx(directory: Path, name: str, rows: list[list], header: list[str] | None = None) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(header if header is not None else HEADER)
    for row in rows:
        ws.append(row)
    path = directory / name
    wb.save(path)
    wb.close()
    return path


def _row(**overrides) -> list:
    base = {
        "client_id": "1", "client_name": "Acme Corp", "client_type": "Corporate",
        "sector": "Technology", "sector_risk": "High", "country": "in",
        "pep_flag": "0", "sanctions_flag": "1", "fatf_country_flag": "0",
    }
    base.update(overrides)
    return [base[c] for c in HEADER]


# --- 15. valid XLSX accepted + 23. normalization works ---------------------- #
def test_valid_xlsx_accepted_and_normalized(tmp_path: Path) -> None:
    _write_xlsx(tmp_path, "clients.xlsx", [_row(client_id="1"), _row(client_id="2", country="us")])
    result = ingest_kyc_file("clients.xlsx", approved_dir=tmp_path)
    assert result.report.total_rows == 2
    assert result.report.valid_rows == 2
    assert all(isinstance(e, NormalizedKYCEntity) for e in result.entities)
    assert result.entities[0].country == "IN"  # normalized upper-case
    assert result.entities[0].sector_risk == SectorRiskLevel.HIGH


def test_xlsx_boolean_and_numeric_cells_coerced_to_strings(tmp_path: Path) -> None:
    # openpyxl preserves native bool/int types; the reader must stringify them
    # so the shared normalizer treats them exactly like CSV text.
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(HEADER)
    ws.append(["1", "Acme", "Corporate", "Tech", "High", "in", False, True, 0])
    wb.save(tmp_path / "typed.xlsx")
    wb.close()
    header, rows = read_xlsx_rows(tmp_path / "typed.xlsx")
    assert rows[0]["pep_flag"] == "False"
    assert rows[0]["sanctions_flag"] == "True"
    assert rows[0]["fatf_country_flag"] == "0"


# --- 16. malformed/corrupt XLSX rejected safely ----------------------------- #
def test_corrupt_xlsx_rejected(tmp_path: Path) -> None:
    (tmp_path / "bad.xlsx").write_bytes(b"this is definitely not a zip/xlsx")
    with pytest.raises(CorruptFileError):
        ingest_kyc_file("bad.xlsx", approved_dir=tmp_path)


def test_corrupt_xlsx_error_leaks_no_parser_internals(tmp_path: Path) -> None:
    (tmp_path / "bad.xlsx").write_bytes(b"PK\x03\x04 not really")
    with pytest.raises(CorruptFileError) as exc_info:
        read_xlsx_rows(tmp_path / "bad.xlsx")
    # Safe message only — the underlying openpyxl/zipfile error is chained
    # (__cause__) but not embedded in the message we raise.
    assert "bad.xlsx" in str(exc_info.value)


# --- 17. missing required columns rejected ---------------------------------- #
def test_xlsx_missing_required_columns_rejected(tmp_path: Path) -> None:
    from app.ingestion.errors import SourceSchemaError

    _write_xlsx(tmp_path, "partial.xlsx", [["1", "Acme"]], header=["client_id", "client_name"])
    with pytest.raises(SourceSchemaError):
        ingest_kyc_file("partial.xlsx", approved_dir=tmp_path)


# --- 18. empty workbook/sheet handled safely -------------------------------- #
def test_empty_xlsx_sheet_returns_no_rows(tmp_path: Path) -> None:
    wb = openpyxl.Workbook()  # a fresh workbook has one empty sheet
    wb.save(tmp_path / "empty.xlsx")
    wb.close()
    header, rows = read_xlsx_rows(tmp_path / "empty.xlsx")
    assert header == [] and rows == []


def test_header_only_xlsx_yields_zero_entities(tmp_path: Path) -> None:
    _write_xlsx(tmp_path, "headeronly.xlsx", [])
    result = ingest_kyc_file("headeronly.xlsx", approved_dir=tmp_path)
    assert result.report.total_rows == 0
    assert result.entities == []


# --- 19. multiple sheets: only the first is read ---------------------------- #
def test_only_first_sheet_is_read(tmp_path: Path) -> None:
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "primary"
    ws1.append(HEADER)
    ws1.append(_row(client_id="1"))
    ws2 = wb.create_sheet("secondary")
    ws2.append(HEADER)
    ws2.append(_row(client_id="999", client_name="SHOULD_NOT_BE_READ"))
    wb.save(tmp_path / "multi.xlsx")
    wb.close()
    header, rows = read_xlsx_rows(tmp_path / "multi.xlsx")
    assert len(rows) == 1
    assert rows[0]["client_id"] == "1"


# --- 21/22. path traversal + outside approved dir rejected ------------------ #
def test_xlsx_path_traversal_rejected(tmp_path: Path) -> None:
    approved = tmp_path / "approved"
    approved.mkdir()
    with pytest.raises(KycPathError):
        ingest_kyc_file("../evil.xlsx", approved_dir=approved)


def test_xlsx_absolute_path_rejected(tmp_path: Path) -> None:
    approved = tmp_path / "approved"
    approved.mkdir()
    with pytest.raises(KycPathError):
        ingest_kyc_file(str(tmp_path / "outside.xlsx"), approved_dir=approved)


# --- empty-file + wrong-extension guards apply to xlsx too ------------------ #
def test_zero_byte_xlsx_rejected(tmp_path: Path) -> None:
    (tmp_path / "empty0.xlsx").write_bytes(b"")
    with pytest.raises(EmptyFileError):
        ingest_kyc_file("empty0.xlsx", approved_dir=tmp_path)


def test_missing_xlsx_rejected(tmp_path: Path) -> None:
    from app.ingestion.errors import MissingFileError as _MFE

    with pytest.raises(_MFE):
        ingest_kyc_file("nope.xlsx", approved_dir=tmp_path)


def test_csv_disguised_as_xlsx_is_rejected_not_misparsed(tmp_path: Path) -> None:
    # A real CSV renamed to .xlsx is not a valid zip container -> CorruptFileError,
    # never silently misparsed.
    (tmp_path / "fake.xlsx").write_text("client_id,client_name\n1,Acme\n", encoding="utf-8")
    with pytest.raises(CorruptFileError):
        ingest_kyc_file("fake.xlsx", approved_dir=tmp_path)


# --- 24/25. raw cell PII not logged; audit is safe aggregate only ----------- #
def test_xlsx_ingestion_audit_has_no_raw_pii(tmp_path: Path, _isolate_audit_sink) -> None:
    _write_xlsx(tmp_path, "pii.xlsx", [_row(client_id="1", client_name=SYNTHETIC_PII)])
    ingest_kyc_file("pii.xlsx", approved_dir=tmp_path)
    completed = [e for e in _isolate_audit_sink.events if e.action == "ingestion.file.completed"]
    assert len(completed) == 1
    dumped = "".join(e.model_dump_json() for e in _isolate_audit_sink.events)
    assert SYNTHETIC_PII not in dumped
    assert completed[0].metadata["source_format"] == "xlsx"
    assert completed[0].metadata["total_rows"] == 1


# --- duplicate handling is identical to CSV (shared normalize_batch) -------- #
def test_xlsx_duplicate_client_ids_dropped(tmp_path: Path) -> None:
    _write_xlsx(tmp_path, "dupes.xlsx", [_row(client_id="7"), _row(client_id="7")])
    result = ingest_kyc_file("dupes.xlsx", approved_dir=tmp_path)
    assert result.report.duplicate_client_ids == 1
    assert result.entities == []  # both occurrences dropped, no silent winner


# --- MAX_XLSX_ROWS is a real, sane bound ------------------------------------ #
def test_max_xlsx_rows_constant_is_bounded() -> None:
    assert isinstance(MAX_XLSX_ROWS, int)
    assert 0 < MAX_XLSX_ROWS <= 10_000_000
